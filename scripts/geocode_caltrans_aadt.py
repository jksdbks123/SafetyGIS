"""
geocode_caltrans_aadt.py
------------------------
Converts Caltrans AADT postmile records to GPS coordinates using the
Caltrans Postmile Query Tool web service, then outputs merged GeoJSON + CSV.

Data sources read:
  data/CaltransAADT/2023-traffic-volumes.xlsx   → mainline AADT (back/ahead)
  data/CaltransAADT/2023-truck-aadt-a11y.xlsx   → truck AADT per postmile
  data/CaltransAADT/Ramp/*.xlsx                 → ramp ADT by district

Outputs (written to data/CaltransAADT/):
  postmile_coords_cache.json   — resumable cache of postmile → (lon, lat)
  aadt_geocoded.csv            — all records with lon/lat columns appended
  aadt_geocoded.geojson        — GeoJSON FeatureCollection (point per record)

Usage:
  python scripts/geocode_caltrans_aadt.py [--counties SAC,FRE] [--rate 2.0]

Options:
  --counties  Comma-separated list of Caltrans 2-3 letter county codes to
              restrict geocoding (e.g. SAC,FRE,HUM). Default: all counties.
  --rate      API requests per second. Default: 2.0 (safe for shared server).
  --no-ramps  Skip ramp files (speeds up processing if only mainline needed).
"""

import argparse
import glob
import json
import logging
import os
import sys
import time
from pathlib import Path

import openpyxl
import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
AADT_DIR = BASE_DIR / "data" / "CaltransAADT"
CACHE_FILE = AADT_DIR / "postmile_coords_cache.json"
OUT_CSV = AADT_DIR / "aadt_geocoded.csv"
OUT_GEOJSON = AADT_DIR / "aadt_geocoded.geojson"

# ---------------------------------------------------------------------------
# Caltrans LRS API
# ---------------------------------------------------------------------------
API_URL = "https://postmile.dot.ca.gov/PMQT/proxy.php"
API_HEADERS = {
    "Content-Type": "application/x-www-form-urlencoded",
    "User-Agent": "SafetyGIS-AADT-Geocoder/1.0",
    "Referer": "https://postmile.dot.ca.gov/PMQT/PostmileQueryTool.html",
}


def _clean(val):
    """Return empty string for None/whitespace values from xlsx."""
    if val is None:
        return ""
    s = str(val).strip()
    return s if s else ""


def _pm_key(county, route, route_sfx, pm_pfx, pm_val, pm_sfx):
    """Canonical cache key for one postmile record."""
    return f"{county}|{route}|{_clean(route_sfx)}|{_clean(pm_pfx)}|{pm_val}|{_clean(pm_sfx)}"


def query_coords(county, route, route_sfx, pm_pfx, pm_val, pm_sfx,
                 session, retries=2):
    """
    Call getCoordinatesForPostmile and return (lon, lat) or None on failure.

    The API expects:
      POST proxy.php
      Content-Type: application/x-www-form-urlencoded
      Body: request=<JSON>

    Response geometry.x = longitude (WGS84), geometry.y = latitude.
    """
    payload = {
        "callType": "getCoordinatesForPostmile",
        "content": {
            "pm": {
                "routeNumber": int(route),
                "routeSuffixCode": _clean(route_sfx),
                "countyCode": county,
                "postmilePrefixCode": _clean(pm_pfx),
                "postmileSuffixCode": _clean(pm_sfx),
                "postmileValue": float(pm_val),
                "alignmentCode": "",
            }
        },
    }
    for attempt in range(retries + 1):
        try:
            r = session.post(
                API_URL,
                data={"request": json.dumps(payload)},
                headers=API_HEADERS,
                timeout=20,
            )
            if r.status_code == 429:
                log.warning("Rate limited (429), sleeping 10 s")
                time.sleep(10)
                continue
            if r.status_code != 200 or not r.text:
                return None
            data = r.json()
            loc = data.get("locations", [{}])[0]
            if loc.get("status") != "esriLocatingOK":
                return None
            geom = loc.get("geometry", {})
            x, y = geom.get("x"), geom.get("y")
            if x is None or y is None:
                return None
            return (round(float(x), 7), round(float(y), 7))
        except Exception as exc:
            log.debug("API error (attempt %d): %s", attempt + 1, exc)
            time.sleep(2)
    return None


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def read_traffic_volumes(path, county_filter=None):
    """
    Read 2023-traffic-volumes.xlsx.
    Returns list of dicts with keys:
      district, route, route_sfx, county, pm_pfx, pm, pm_sfx,
      description, back_peak_hour, back_peak_madt, back_aadt,
      ahead_peak_hour, ahead_peak_madt, ahead_aadt
    """
    records = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["2023 AADT DATA"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        district, route, route_sfx, county = row[0], row[1], row[2], row[3]
        pm_pfx, pm, pm_sfx = row[4], row[5], row[6]
        if pm is None:
            continue
        if county_filter and _clean(county).upper() not in county_filter:
            continue
        records.append({
            "source": "mainline",
            "district": _clean(district),
            "route": _clean(route).lstrip("0") or "0",
            "route_sfx": _clean(route_sfx),
            "county": _clean(county).upper(),
            "pm_pfx": _clean(pm_pfx),
            "pm": float(pm),
            "pm_sfx": _clean(pm_sfx),
            "description": _clean(row[7]),
            "back_peak_hour": _clean(row[8]),
            "back_peak_madt": _clean(row[9]),
            "back_aadt": _clean(row[10]),
            "ahead_peak_hour": _clean(row[11]),
            "ahead_peak_madt": _clean(row[12]),
            "ahead_aadt": _clean(row[13]),
            # Derive a single representative AADT (prefer ahead, fall back to back)
            "aadt": _clean(row[13]) or _clean(row[10]),
        })
    wb.close()
    return records


def read_truck_aadt(path, county_filter=None):
    """Read 2023-truck-aadt-a11y.xlsx."""
    records = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Truck AADT 2023 "]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        route, route_sfx, district, county = row[0], row[1], row[2], row[3]
        pm_pfx, pm, pm_sfx = row[4], row[5], row[6]
        if pm is None:
            continue
        if county_filter and _clean(county).upper() not in county_filter:
            continue
        records.append({
            "source": "truck",
            "district": _clean(district),
            "route": _clean(route).lstrip("0") or "0",
            "route_sfx": _clean(route_sfx),
            "county": _clean(county).upper(),
            "pm_pfx": _clean(pm_pfx),
            "pm": float(pm),
            "pm_sfx": _clean(pm_sfx),
            "description": _clean(row[8]),
            "vehicle_aadt": _clean(row[9]),
            "truck_aadt": _clean(row[10]),
            "truck_pct": _clean(row[11]),
            "aadt": _clean(row[9]),  # total vehicle AADT
        })
    wb.close()
    return records


def read_ramp_file(path, county_filter=None):
    """Read one district ramp ADT xlsx."""
    records = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Last sheet is the data sheet
    sheet_name = wb.sheetnames[-1]
    ws = wb[sheet_name]
    header = None
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        if header is None:
            header = [str(c).strip().upper() if c else "" for c in row]
            continue
        if row[3] is None:  # CNTY column
            continue
        # Columns: YRS, DIST_COUNTY_ROUTE, DIST, CNTY, RTE, RTE_SFX,
        #          PM_PFX, POSTMILE, PM_SFX, LOCATION_DESC,
        #          YR_2014..YR_2023
        county = _clean(row[3]).upper()
        if county_filter and county not in county_filter:
            continue
        route = _clean(row[4]).lstrip("0") or "0"
        route_sfx = _clean(row[5])
        pm_pfx = _clean(row[6])
        pm = row[7]
        pm_sfx = _clean(row[8])
        if pm is None:
            continue
        # Most recent non-null ADT across YR_2014..YR_2023 (columns 10..19)
        aadt_val = ""
        for v in reversed(row[10:20]):
            if v is not None:
                aadt_val = str(int(v))
                break
        records.append({
            "source": "ramp",
            "district": _clean(row[2]),
            "route": route,
            "route_sfx": route_sfx,
            "county": county,
            "pm_pfx": pm_pfx,
            "pm": float(pm),
            "pm_sfx": pm_sfx,
            "description": _clean(row[9]),
            "aadt": aadt_val,
        })
    wb.close()
    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Geocode Caltrans AADT postmiles to GPS")
    parser.add_argument("--counties", default="", help="Comma-separated county codes, e.g. SAC,FRE")
    parser.add_argument("--rate", type=float, default=2.0, help="API requests per second (default 2.0)")
    parser.add_argument("--no-ramps", action="store_true", help="Skip ramp files")
    args = parser.parse_args()

    county_filter = None
    if args.counties:
        county_filter = {c.strip().upper() for c in args.counties.split(",")}
        log.info("Filtering to counties: %s", county_filter)

    sleep_interval = 1.0 / max(args.rate, 0.1)

    # ------------------------------------------------------------------
    # 1. Load all records
    # ------------------------------------------------------------------
    log.info("Reading traffic volumes...")
    records = read_traffic_volumes(AADT_DIR / "2023-traffic-volumes.xlsx", county_filter)
    log.info("  %d mainline records", len(records))

    log.info("Reading truck AADT...")
    truck_records = read_truck_aadt(AADT_DIR / "2023-truck-aadt-a11y.xlsx", county_filter)
    log.info("  %d truck records", len(truck_records))
    records.extend(truck_records)

    if not args.no_ramps:
        log.info("Reading ramp files...")
        ramp_files = sorted(glob.glob(str(AADT_DIR / "Ramp" / "*.xlsx")))
        for ramp_path in ramp_files:
            ramp_records = read_ramp_file(ramp_path, county_filter)
            records.extend(ramp_records)
        log.info("  %d total records after ramps", len(records))

    # ------------------------------------------------------------------
    # 2. Build deduplicated postmile set
    # ------------------------------------------------------------------
    pm_keys = {}
    for rec in records:
        key = _pm_key(
            rec["county"], rec["route"], rec["route_sfx"],
            rec["pm_pfx"], rec["pm"], rec["pm_sfx"]
        )
        pm_keys[key] = rec  # just need one representative record per key

    log.info("Unique postmiles to geocode: %d", len(pm_keys))

    # ------------------------------------------------------------------
    # 3. Load coordinate cache (resume support)
    # ------------------------------------------------------------------
    cache = {}
    if CACHE_FILE.exists():
        with open(CACHE_FILE, "r") as f:
            cache = json.load(f)
        log.info("Loaded %d cached coordinates", len(cache))

    # ------------------------------------------------------------------
    # 4. Query API for uncached postmiles
    # ------------------------------------------------------------------
    session = requests.Session()
    uncached = [k for k in pm_keys if k not in cache]
    log.info("Need to query %d postmiles (%.0f min at %.1f req/s)",
             len(uncached), len(uncached) / args.rate / 60, args.rate)

    for i, key in enumerate(uncached):
        rec = pm_keys[key]
        coords = query_coords(
            rec["county"], rec["route"], rec["route_sfx"],
            rec["pm_pfx"], rec["pm"], rec["pm_sfx"],
            session,
        )
        # Cache even failed lookups as null to avoid re-querying
        cache[key] = coords  # (lon, lat) or None

        if (i + 1) % 100 == 0 or (i + 1) == len(uncached):
            with open(CACHE_FILE, "w") as f:
                json.dump(cache, f)
            ok = sum(1 for v in cache.values() if v is not None)
            log.info("Progress: %d/%d queried | %d with coords | cache saved",
                     i + 1, len(uncached), ok)

        time.sleep(sleep_interval)

    # ------------------------------------------------------------------
    # 5. Merge coordinates back into all records
    # ------------------------------------------------------------------
    log.info("Merging coordinates into records...")
    geocoded = []
    failed = 0
    for rec in records:
        key = _pm_key(
            rec["county"], rec["route"], rec["route_sfx"],
            rec["pm_pfx"], rec["pm"], rec["pm_sfx"]
        )
        coords = cache.get(key)
        if coords:
            rec["lon"], rec["lat"] = coords
            geocoded.append(rec)
        else:
            failed += 1

    log.info("Geocoded: %d | Failed/no coords: %d", len(geocoded), failed)

    # ------------------------------------------------------------------
    # 6. Write CSV
    # ------------------------------------------------------------------
    import csv

    # Collect all unique field names across all record types
    all_fields = ["source", "district", "county", "route", "route_sfx",
                  "pm_pfx", "pm", "pm_sfx", "description", "aadt",
                  "back_aadt", "ahead_aadt", "back_peak_hour", "back_peak_madt",
                  "ahead_peak_hour", "ahead_peak_madt",
                  "vehicle_aadt", "truck_aadt", "truck_pct",
                  "lon", "lat"]

    log.info("Writing CSV to %s", OUT_CSV)
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(geocoded)

    # ------------------------------------------------------------------
    # 7. Write GeoJSON
    # ------------------------------------------------------------------
    log.info("Writing GeoJSON to %s", OUT_GEOJSON)
    features = []
    for rec in geocoded:
        props = {k: v for k, v in rec.items() if k not in ("lon", "lat")}
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [rec["lon"], rec["lat"]]},
            "properties": props,
        })

    fc = {"type": "FeatureCollection", "features": features}
    with open(OUT_GEOJSON, "w", encoding="utf-8") as f:
        json.dump(fc, f, separators=(",", ":"))

    log.info("Done. %d features written.", len(features))

    # Summary by source
    from collections import Counter
    by_source = Counter(r["source"] for r in geocoded)
    for src, cnt in sorted(by_source.items()):
        log.info("  %-10s %d records", src, cnt)


if __name__ == "__main__":
    main()
